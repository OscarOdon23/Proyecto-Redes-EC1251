import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import cmath
# Nombre del archivo original y la hoja de datos de f_and_output
filename = "data_io.xlsx"
sheetname = "f_and_ouput"

# Cargar el libro de trabajo de Excel
workbook = openpyxl.load_workbook(filename)

# Obtener la hoja de datos de f_and_output
worksheet = workbook[sheetname]

# Obtener el nombre del archivo de salida de la celda B2
output_filename = worksheet.cell(row=2, column=2).value

#--------------------------------------------------------Lineamientos---------------------------------------------------------------------------
np.set_printoptions(precision=4, suppress=True, formatter={"complexfloat": "{:.4f}".format}) #linea opcional de formato
#Lectura de Excel
df_f = pd.read_excel("data_io.xlsx", sheet_name="f_and_ouput", usecols=[0, 1], header=None, nrows=2)
freq_value = float(df_f.iloc[0, 1])
output_file_name = str(df_f.iloc[1, 1])
df_V_f = pd.read_excel("data_io.xlsx", 1 )
df_I_f = pd.read_excel("data_io.xlsx", 2 )
df_Z= pd.read_excel("data_io.xlsx", 3, )

# Define las hojas y las columnas que deseas excluir
excluded_sheets_columns = {
    "V_fuente": ['B'],  # Excluye la segunda columna en la hoja 'V_fuente'
    "I_fuente": ['B'],  # Excluye la segunda columna en la hoja 'I_fuente'
    "Z": ['C']  # Excluye la tercera columna en la hoja 'Z'
}

# Procesa solo las hojas especificadas
for sheet_name, excluded_columns in excluded_sheets_columns.items():
    sheet = workbook[sheet_name]
    
    # Encuentra la última fila con datos en cualquier columna que no sea una columna excluida
    last_row_with_data = max((cell.row for row in sheet.iter_rows() for cell in row if cell.value is not None and openpyxl.utils.cell.get_column_letter(cell.column) not in excluded_columns), default=0)

    for row in sheet.iter_rows(max_row=last_row_with_data):
        for cell in row:
            column_letter = openpyxl.utils.cell.get_column_letter(cell.column)
            if column_letter not in excluded_columns and cell.value is None:
                cell.value = 0

workbook.save(filename=output_filename)

#Relleno de ceros
V_f=df_V_f.fillna({"Rf (ohms)":0,"Lf (mH)":0,"Cf (uF)":0})
I_f=df_I_f.fillna({"Rf (ohms)":0,"Lf (mH)":0,"Cf (uF)":0})
Z=df_Z.fillna({"R (ohms)":0,"L (mH)":0,"C (uF)":0})

# Leer datos de V_fuente y vectores de la fuente
bus_i_v = np.array(V_f.iloc[:,0])
vpico = np.array(V_f.iloc[:,2])
Rf_v = np.array(V_f.iloc[:, 4])
Lf_v = np.array(V_f.iloc[:, 5])
Cf_v = np.array(V_f.iloc[:, 6])

# Crear una lista para almacenar los warnings
warnings = []

# Validar los datos de V_fuente y publicar el warning en el archivo de salida
for i in range(len(bus_i_v)):
    try:
        if vpico[i] <= 0:
            raise ValueError(f"Introdujo un valor negativo en la fila {i+2}")

        if Rf_v[i] < 0:
            raise ValueError(f"Introdujo un valor negativo en la fila {i+2}")

        if Lf_v[i] < 0:
            raise ValueError(f"Introdujo un valor negativo en la fila {i+2}")

        if Cf_v[i] < 0:
            raise ValueError(f"Introdujo un valor negativo en la fila {i+2}")

    except ValueError as e:
        worksheet_v = workbook["V_fuente"]
        worksheet_v.cell(row=i+2, column=2, value=str(e))
        
        warnings.append(str(e))

# Leer datos de I_fuente y vectores de la fuente
bus_i_i = np.array(I_f.iloc[:, 0])
ipico = np.array(I_f.iloc[:, 2])
t0_i = np.array(I_f.iloc[:, 3])
Rf_i = np.array(I_f.iloc[:, 4])
Lf_i = np.array(I_f.iloc[:, 5])
Cf_i = np.array(I_f.iloc[:, 6])

# Validar los datos de I_fuente
for i in range(len(bus_i_i)):
    try:
        if ipico[i] <= 0:
            raise ValueError(f"Introdujo un valor negativo en la fila {i+2}")

        if Rf_i[i] < 0:
            raise ValueError(f"Introdujo un valor negativo en la fila {i+2}")

        if Lf_i[i] < 0:
            raise ValueError(f"Introdujo un valor negativo en la fila {i+2}")

        if Cf_i[i] < 0:
            raise ValueError(f"Introdujo un valor negativo en la fila {i+2}")

    except ValueError as e:
        worksheet_i = workbook["I_fuente"]
        worksheet_i.cell(row=i+2, column=2, value=str(e))
        
        warnings.append(str(e))

# Leer datos de Z y vectores de linea
bline_i = np.array(Z.iloc[:, 0])
bline_j = np.array(Z.iloc[:, 1])
Rline = np.array(Z.iloc[:, 3])
Lline = np.array(Z.iloc[:, 4])
Cline = np.array(Z.iloc[:, 5])

# Validar los datos de Z y publicar el warning en el archivo de salida
for i in range(len(bline_i)):
    try:
        if Rline[i] < 0:
            raise ValueError(f"Introdujo un valor negativo en la fila {i+2}")

        if Lline[i] < 0:
            raise ValueError(f"Introdujo un valor negativo en la fila {i+2}")

        if Cline[i] < 0:
            raise ValueError(f"Introdujo un valor negativo en la fila {i+2}")

    except ValueError as e:
        worksheet_z = workbook["Z"]
        worksheet_z.cell(row=i+2, column=3, value=str(e))
        
        warnings.append(str(e))

# Si se encontraron warnings, imprimirlos y detener el programa
if warnings:
    print("Se encontraron los siguientes warnings:")
    for warning in warnings:
        print(warning)
    raise Exception("Se encontraron valores negativos. Por favor, corregir e intentar de nuevo.")

workbook.save(filename=output_filename)
#-------------------------------------------------------------------------Operaciones---------------------------------------------------------------
#Calculos de reactancias, impedancias y admitancias

#inductivas
xline_l=2*np.pi*freq_value*1E-3*Lline*1j
xf_v_l=2*np.pi*freq_value*1E-3*Lf_v*1j
xf_i_l=2*np.pi*freq_value*1E-3*Lf_i*1j
#capacitivas
xline_c = np.zeros(len(Cline),dtype=complex)
for i in range(len(Cline)):
    if Cline[i]!=0:
        xline_c[i]= (1j)/(2*np.pi*freq_value*1E-6*Cline[i])
    else:
        continue
xf_v_c = np.zeros(len(Cf_v),dtype=complex)
for i in range(len(Cf_v)):
    if Cf_v[i]!=0:
        xf_v_c[i]= (1j)/(2*np.pi*freq_value*1E-6*Cf_v[i])
    else:
        continue
xf_i_c = np.zeros(len(Cf_i),dtype=complex)
for i in range(len(Cf_i)):
    if Cf_i[i]!=0:
        xf_i_c[i]= (1j)/(2*np.pi*freq_value*1E-6*Cf_i[i])
    else:
        continue
#Reactancia total
xline=(xline_l - xline_c)
xf_i= (xf_i_l - xf_i_c)

xf_v=(xf_v_l-xf_v_c)
#impedancia total
zline= Rline + xline


zf_v= Rf_v +xf_v

#Matriz ordenada segun el nodo
arrays = [bus_i_v, bline_j, bline_i, bus_i_i]
max_values = []

for arr in arrays:
    if arr.size != 0:  # Comprueba si el array está vacío
        max_values.append(np.max(arr))

if max_values:  # Comprueba si la lista de valores máximos está vacía
    total_nodos = max(max_values)
else:
    print("Todos los arrays están vacíos.")

zf_v2 = [0]*total_nodos
for i in range(len(bus_i_v)):
    node = bus_i_v[i] - 1  
    if i < len(Rf_v) and i < len(xf_v):
        zf_v2[node] = Rf_v[i] + xf_v[i]
    elif i < len(Rf_v):
        zf_v2[node] = Rf_v[i]
    elif i < len(xf_v):
        zf_v2[node] = xf_v[i]

zf_i= Rf_i + xf_i

zf_i2 = [0]*total_nodos
for i in range(len(bus_i_i)):
    node = bus_i_i[i] - 1  
    if i < len(Rf_i) and i < len(xf_i):
        zf_i2[node] = Rf_i[i] + xf_i[i]
    elif i < len(Rf_i):
        zf_i2[node] = Rf_i[i]
    elif i < len(xf_i):
        zf_v2[node] = xf_i[i]

# admitancia total
yline = np.zeros(len(zline),dtype=complex)
for i in range(len(zline)):
    if zline[i]!= 0:
        yline[i] = (1/zline[i])
    else:
        continue

yf_v = np.zeros(len(zf_v),dtype=complex)
for i in range(len(zf_v)):
    if zf_v[i]!= 0:
        yf_v[i] = (1/zf_v[i])
    else:
        continue

yf_i = np.zeros(len(zf_i),dtype=complex)
for i in range(len(zf_i)):
    if zf_i[i]!= 0:
        yf_i[i] = (1/zf_i[i])
    else:
        continue    
    
#-------------------------------------------------------YBUS-----------------------------------------------------------------------------    
#Contrucción de la matriz de admitancia de línea
max_valor= np.max([bline_i, bline_j])
yline_out= np.zeros(((max_valor, max_valor)),dtype=complex)
for k in range(len(bline_i)):
    i=int(bline_i[k]-1)
    j=int(bline_j[k]-1)
    if i== -1 or j==-1:
        continue
    else:
        yline_out[i,j]= -1*yline[k]
        yline_out[j,i]= yline_out[i,j]

#Admitancias de la diagonal
diagonal = yline_out.sum(axis=1)
diagonal = np.diag(diagonal)
yline_out = yline_out - diagonal
for k in range(len(bline_i)):
    i = int(bline_i[k]-1)
    j = int(bline_j[k]-1)
    if i == -1 or j == -1:
        if i == -1:
            yline_out[j,j] = yline_out[j,j] + yline[k]
        elif j == -1:
            yline_out[i,i] = yline_out[i,i] + yline[k]

for k in range(len(bus_i_v)):
    i = int(bus_i_v[k]-1)
    if i != -1:
        yline_out[i,i] += yf_v[k]

for k in range(len(bus_i_i)):
    i = int(bus_i_i[k]-1)
    if i != -1:
        yline_out[i,i] += yf_i[k]

#-----------------------------------------------------------------TRANSFORMACION DE FUENTES-------------------------------------------------------------------------------------------------------
#Matriz de voltaje de fuente en rectangulares
phase_shift_rad = np.deg2rad(df_V_f['Corriemento  de onda en t(seg)'])
phase_shift_radI = np.deg2rad(df_I_f['Corriemento to (seg)'])

vpico_complexmax = np.zeros(total_nodos, dtype=complex)
for i in range(len(df_V_f)):
    nodo = int(df_V_f.iloc[i]['Bus i'] - 1)
    vpico_complexmax[nodo] = cmath.rect(df_V_f['Vpico f (V)'][i]/np.sqrt(2), phase_shift_rad[i])

Ipico_complexmax = np.zeros(total_nodos, dtype=complex)
for i in range(len(df_I_f)):
    nodo = int(df_I_f.iloc[i]['Bus i'] - 1)
    Ipico_complexmax[nodo] = cmath.rect(df_I_f['I pico f (A)'][i]/np.sqrt(2), phase_shift_radI[i])

ipicof = np.zeros(total_nodos, dtype=complex)
for i in range(total_nodos):
    if Ipico_complexmax[i] == 0 and vpico_complexmax[i] != 0:
        if zf_v2[i] == 0:
            print(f"Warning: La fuente de voltaje en el nodo {i+1} no puede ser convertida a una fuente de corriente porque su impedancia es cero.")
        else:
            ipicof[i] = vpico_complexmax[i] / zf_v2[i]

    elif Ipico_complexmax[i] != 0 and vpico_complexmax[i] == 0:
        ipicof[i] = Ipico_complexmax[i]

    elif Ipico_complexmax[i] != 0 and vpico_complexmax[i] != 0:
        if zf_v2[i] == 0:
            print(f"Warning: La fuente de voltaje en el nodo {i+1} no puede ser convertida a una fuente de corriente porque su impedancia es cero.")
        else:
            ipicof[i] = (vpico_complexmax[i] / zf_v2[i]) + Ipico_complexmax[i]

ipicof_phasor=[cmath.polar(i) for i in ipicof]
ipicof_phasor_degrees = np.array([complex(round(magnitude, 4), round(np.rad2deg(phase), 4)) for magnitude, phase in ipicof_phasor])
vnodo= np.linalg.solve(yline_out,ipicof)

vnodo_phasor=[cmath.polar(i) for i in vnodo]
vnodo_phasor_degrees = np.array([complex(round(magnitude, 4), round(np.rad2deg(phase), 4)) for magnitude, phase in vnodo_phasor])
#------------------------------------------------------------Transformacion a fuentes de voltaje----------------------------------------------------
vpicof = np.zeros(total_nodos, dtype=complex)
vpicoitv = np.zeros(total_nodos, dtype=complex)

for i in range(total_nodos):
    if Ipico_complexmax[i] != 0 and vpico_complexmax[i] == 0:
        if zf_i2[i] == 0:
            print(f"Warning: La fuente de corriente en el nodo {i+1} no puede ser convertida a una fuente de voltaje porque su impedancia es cero.")
        else:
            vpicof[i] = Ipico_complexmax[i] * zf_i2[i]

    elif Ipico_complexmax[i] == 0 and vpico_complexmax[i] != 0:
        vpicof[i] = vpico_complexmax[i]

    elif Ipico_complexmax[i] != 0 and vpico_complexmax[i] != 0:
        if zf_i2[i] == 0:
            print(f"Warning: La fuente de corriente en el nodo {i+1} no puede ser convertida a una fuente de voltaje porque su impedancia es cero.")
        else:
            vpicoitv[i] = (Ipico_complexmax[i] * zf_i2[i])
            vpicof= vpicoitv[i]+vpico_complexmax[i]

#-----------------------------------------------------------------Vth_AND_ZTH------------------------------------------------------------

# Cálculo de los valores
yline_inv = np.linalg.inv(yline_out)
zth = np.diagonal(yline_inv)
Rth = zth.real
Qth = zth.imag
Vth=(vnodo_phasor_degrees.real)/1000
AngleVth=vnodo_phasor_degrees.imag

df_Rth = pd.DataFrame(Rth, columns=['Rth'])
df_Qth = pd.DataFrame(Qth, columns=['Qth'])
df_Vth = pd.DataFrame(Vth, columns=['Vth'])
df_AngleVth = pd.DataFrame(AngleVth, columns=['AngleVth'])

worksheet = workbook['VTH_AND_ZTH']

for bus_index, (value_Rth, value_Qth, value_Vth, value_AngleVth) in enumerate(zip(df_Rth['Rth'], df_Qth['Qth'], df_Vth['Vth'], df_AngleVth['AngleVth']), start=1):
    worksheet.cell(row=bus_index + 1, column=2, value=value_Vth)
    worksheet.cell(row=bus_index + 1, column=3, value=value_AngleVth)
    worksheet.cell(row=bus_index + 1, column=4, value=value_Rth)
    worksheet.cell(row=bus_index + 1, column=5, value=value_Qth)

workbook.save(filename=output_filename)

#---------------------------------------------------------------------SFuente-------------------------------------------------------------------------------------------------------------------

# Caso 1: Solo hay fuentes de voltaje
if len(vpico_complexmax) > 0 and len(vpicoitv) == 0:
    Vdif = (vpico_complexmax - vnodo)
    Vdif = np.array(Vdif, dtype=complex)
    zf_v2 = np.array(zf_v2, dtype=complex)

    Idif = np.zeros_like(Vdif)
    mask = zf_v2 != 0
    Idif[mask] = Vdif[mask] / zf_v2[mask]

    Sf = (vnodo * np.conj(Idif))

    Pf = Sf.real
    Qf = Sf.imag

# Caso 2: Solo hay fuentes de corriente transformadas en fuentes de voltaje
elif len(vpicoitv) > 0 and len(vpico_complexmax) == 0:
    Vdif = (vpicoitv - vnodo)
    Vdif = np.array(Vdif, dtype=complex)
    zf_i2 = np.array(zf_i2, dtype=complex)

    Idif = np.zeros_like(Vdif)
    mask = zf_i2 != 0
    Idif[mask] = Vdif[mask] / zf_i2[mask]

    Sf = (vnodo * np.conj(Idif))

    Pf = Sf.real
    Qf = Sf.imag

# Caso 3: Existen ambos tipos de fuentes, de voltaje y corriente transformada en voltaje
elif len(vpico_complexmax) > 0 and len(vpicoitv) > 0:
    zf_v2 = np.array(zf_v2, dtype=complex)
    zf_i2 = np.array(zf_i2, dtype=complex)

    Vdif_v = (vpico_complexmax - vnodo)
    Vdif_v = np.array(Vdif_v, dtype=complex)

    Idif_v = np.zeros_like(Vdif_v)
    mask_v = zf_v2 != 0
    Idif_v[mask_v] = Vdif_v[mask_v] / zf_v2[mask_v]

    Vdif_i = (vpicoitv - vnodo)
    Vdif_i = np.array(Vdif_i, dtype=complex)

    Idif_i = np.zeros_like(Vdif_i)
    mask_i = zf_i2 != 0
    Idif_i[mask_i] = Vdif_i[mask_i] / zf_i2[mask_i]

    Vdif_total = Vdif_v + Vdif_i
    Idif_total = Idif_v + Idif_i

    Sf_total = (vnodo * np.conj(Idif_total))

    Pf= Sf_total.real
    Qf= Sf_total.imag

# Combina los arrays
combined = np.concatenate([bline_i, bline_j, bus_i_i])
combined_non_zero = combined[combined != 0]

if combined_non_zero.size == 0:
    print("Todos los arrays están vacíos o sólo contienen ceros.")
    bus_i_vf = np.array([])
else:
    min_value = combined_non_zero.min()
    max_value = combined.max()
    bus_i_vf = np.arange(min_value, max_value + 1)

busjv = np.zeros_like(bus_i_vf)
df_Pf = pd.DataFrame(Pf, columns=['Pf'])
df_Qf = pd.DataFrame(Qf, columns=['Qf'])
df_bus_i_vf = pd.DataFrame(bus_i_vf, columns=['bus_i_v'])
df_busjv = pd.DataFrame(busjv, columns=['busjv'])

worksheet = workbook['Sfuente']

for bus_index, (value_Pf, value_Qf, value_bus_i_vf, value_busjv) in enumerate(zip(df_Pf['Pf'], df_Qf['Qf'], df_bus_i_vf['bus_i_v'], df_busjv['busjv']), start=1):
    worksheet.cell(row=bus_index + 1, column=1, value=value_bus_i_vf)
    worksheet.cell(row=bus_index + 1, column=2, value=value_busjv)
    worksheet.cell(row=bus_index + 1, column=3, value=value_Pf)
    worksheet.cell(row=bus_index + 1, column=4, value=value_Qf)

workbook.save(filename=output_filename)

#--------------------------------------------------------------S_Z----------------------------------------------------------------------------------------------------------------------------

s_z = np.zeros((len(bline_i),1),dtype=complex)
for i in range(len(zline)):
    if bline_i[i] ==0 or  bline_j[i] == 0 :
        if bline_j[i] == 0:
            s_z[i] = (vnodo_phasor_degrees.real[bline_i[i]-1] ** 2) / (np.conj(zline[i]))
        elif bline_i[i] ==0:
            s_z[i] = (vnodo_phasor_degrees.real[bline_j[i]-1] ** 2) / (np.conj(zline[i]))
    elif bline_i[i] != 0 or  bline_j[i] != 0:
        s_z[i] = (((vnodo[bline_i[i]-1] - vnodo[bline_j[i]-1]))*np.conj(((vnodo[bline_i[i]-1] - vnodo[bline_j[i]-1]))))/(np.conj(zline[i]))
                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                         
P_z= s_z.real  
Q_z= s_z.imag

df_bline_i = pd.DataFrame(bline_i, columns=['bline_i'])
df_bline_j = pd.DataFrame(bline_j, columns=['bline_j'])
df_P_z = pd.DataFrame(P_z, columns=['P_z'])
df_Q_z = pd.DataFrame(Q_z, columns=['Q_z'])

worksheet = workbook['S_Z']

for bus_index, (value_bline_i, value_bline_j, value_P_z, value_Q_z) in enumerate(zip(df_bline_i['bline_i'], df_bline_j['bline_j'], df_P_z['P_z'], df_Q_z['Q_z']), start=1):
    worksheet.cell(row=bus_index + 1, column=1, value=value_bline_i)
    worksheet.cell(row=bus_index + 1, column=2, value=value_bline_j)
    worksheet.cell(row=bus_index + 1, column=3, value=value_P_z)
    worksheet.cell(row=bus_index + 1, column=4, value=value_Q_z)

workbook.save(filename=output_filename)

#----------------------------------------------------------------BALANCE_S-----------------------------------------------------------------------------------------------------------

# Calcular los totales
Pf_total = Pf.sum()
Qf_total = Qf.sum()
Pz_total = P_z.sum()
Qz_total = Q_z.sum()

# Calcular los deltas
Delta_P = Pf_total - Pz_total
Delta_Q = Qf_total - Qz_total

df_Pf_total = pd.DataFrame([Pf_total], columns=['Pf_total'])
df_Qf_total = pd.DataFrame([Qf_total], columns=['Qf_total'])
df_Pz_total = pd.DataFrame([Pz_total], columns=['Pz_total'])
df_Qz_total = pd.DataFrame([Qz_total], columns=['Qz_total'])
df_Delta_P = pd.DataFrame([Delta_P], columns=['Delta_P'])
df_Delta_Q = pd.DataFrame([Delta_Q], columns=['Delta_Q'])

worksheet = workbook['Balance_S']

for bus_index, (value_Pf_total, value_Qf_total, value_Pz_total, value_Qz_total, value_Delta_P, value_Delta_Q) in enumerate(zip(df_Pf_total['Pf_total'], df_Qf_total['Qf_total'], df_Pz_total['Pz_total'], df_Qz_total['Qz_total'], df_Delta_P['Delta_P'], df_Delta_Q['Delta_Q']), start=1):
    worksheet.cell(row=bus_index + 1, column=1, value=value_Pf_total)
    worksheet.cell(row=bus_index + 1, column=2, value=value_Qf_total)
    worksheet.cell(row=bus_index + 1, column=3, value=value_Pz_total)
    worksheet.cell(row=bus_index + 1, column=4, value=value_Qz_total)
    worksheet.cell(row=bus_index + 1, column=5, value=value_Delta_P)
    worksheet.cell(row=bus_index + 1, column=6, value=value_Delta_Q)

workbook.save(filename=output_filename)